"""
ELECTRO-ORDINATEUR — Application Flask principale
"""
import os
import json
from datetime import datetime, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, session, redirect, url_for,
    jsonify, flash, g, abort, send_from_directory, Response
)
from flask_login import (
    LoginManager, login_user, logout_user, login_required,
    current_user
)
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from dotenv import load_dotenv
import openpyxl
from io import BytesIO

from config import config
from database.models import db, Product, Category, ProductImage, SiteSettings, Admin, ActivityLog, WhatsAppClick, Sale, SaleItem, Payment
from utils.image_handler import save_product_image, delete_product_image, save_payment_proof
from utils.helpers import (
    get_lang, load_translations, build_whatsapp_url,
    build_whatsapp_group_url, format_price, slugify, log_action
)

load_dotenv()


# ─────────────────────────────────────────────────────────
#  App Factory
# ─────────────────────────────────────────────────────────
def create_app(config_name: str = None) -> Flask:
    if config_name is None:
        config_name = os.getenv("FLASK_ENV", "development")

    app = Flask(__name__)
    app.config.from_object(config.get(config_name, config["default"]))

    # Ensure upload folder exists
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
    os.makedirs(os.path.join(app.root_path, "database"), exist_ok=True)

    # Extensions
    db.init_app(app)
    csrf = CSRFProtect(app)
    login_manager = LoginManager(app)
    login_manager.login_view = "admin_login"
    login_manager.login_message = "Veuillez vous connecter pour accéder à l'administration."
    login_manager.login_message_category = "warning"

    limiter = Limiter(
        get_remote_address,
        app=app,
        default_limits=["500 per day", "100 per hour"],
        storage_uri="memory://",
    )

    load_translations(app)

    # ──────────────────────────────────────────
    #  Context Processors
    # ──────────────────────────────────────────
    @app.context_processor
    def inject_globals():
        lang = session.get("lang", "fr")
        translations = app.config.get("TRANSLATIONS", {})
        t_dict = translations.get(lang, {})

        with app.app_context():
            settings = SiteSettings.get_all_dict()
            categories = Category.query.filter_by(active=True).order_by(Category.ordre).all()

        return dict(
            lang=lang,
            is_rtl=(lang == "ar"),
            t=t_dict,
            settings=settings,
            categories=categories,
            now=datetime.utcnow(),
            format_price=format_price,
        )

    @login_manager.user_loader
    def load_user(user_id):
        return Admin.query.get(int(user_id))

    # ──────────────────────────────────────────
    #  Decorators
    # ──────────────────────────────────────────
    def editor_required(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.can_edit:
                flash("Vous n'avez pas les permissions nécessaires.", "danger")
                return redirect(url_for("admin_dashboard"))
            return f(*args, **kwargs)
        return decorated

    def superadmin_required(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not current_user.is_superadmin:
                flash("Action réservée au Super Administrateur.", "danger")
                return redirect(url_for("admin_dashboard"))
            return f(*args, **kwargs)
        return decorated

    # ──────────────────────────────────────────
    #  Helpers
    # ──────────────────────────────────────────
    def get_base_url():
        return request.host_url.rstrip("/")

    # ═══════════════════════════════════════════
    #  PUBLIC ROUTES
    # ═══════════════════════════════════════════

    @app.route("/")
    def index():
        lang = session.get("lang", "fr")
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", app.config["ITEMS_PER_PAGE"], type=int)
        sort = request.args.get("sort", "newest")
        cat_slug = request.args.get("categorie", "")
        brand = request.args.get("marque", "")
        etat = request.args.get("etat", "")
        prix_min = request.args.get("prix_min", type=float)
        prix_max = request.args.get("prix_max", type=float)
        in_stock_only = request.args.get("in_stock", "")

        query = Product.query.filter_by(statut="visible")

        if cat_slug:
            cat = Category.query.filter_by(slug=cat_slug).first()
            if cat:
                query = query.filter_by(categorie_id=cat.id)

        if brand:
            query = query.filter(Product.marque.ilike(f"%{brand}%"))

        if etat:
            query = query.filter_by(etat=etat)

        if prix_min is not None:
            query = query.filter(Product.prix >= prix_min)
        if prix_max is not None:
            query = query.filter(Product.prix <= prix_max)

        if in_stock_only:
            query = query.filter(Product.quantite_stock > 0)

        if sort == "price_asc":
            query = query.order_by(Product.prix.asc())
        elif sort == "price_desc":
            query = query.order_by(Product.prix.desc())
        elif sort == "popular":
            query = query.order_by(Product.nb_vues.desc())
        else:  # newest
            query = query.order_by(Product.date_creation.desc())

        products_paginated = query.paginate(page=page, per_page=per_page, error_out=False)

        featured = Product.query.filter_by(statut="visible", en_vedette=True).limit(6).all()
        brands = [r[0] for r in db.session.query(Product.marque).filter(Product.marque != "").distinct().all()]

        return render_template(
            "public/index.html",
            products=products_paginated,
            featured=featured,
            brands=brands,
            current_sort=sort,
            current_cat=cat_slug,
            current_brand=brand,
            current_etat=etat,
            current_prix_min=prix_min,
            current_prix_max=prix_max,
            in_stock_only=in_stock_only,
        )

    @app.route("/produit/<int:product_id>")
    def product_detail(product_id):
        product = Product.query.filter_by(id=product_id, statut="visible").first_or_404()
        lang = session.get("lang", "fr")

        # Track view
        product.nb_vues += 1
        db.session.commit()

        wa_url = build_whatsapp_url(product, lang, get_base_url())
        similar = Product.query.filter(
            Product.categorie_id == product.categorie_id,
            Product.id != product.id,
            Product.statut == "visible",
        ).limit(4).all()

        specs = product.get_specs_dict()

        return render_template(
            "public/product.html",
            product=product,
            wa_url=wa_url,
            similar=similar,
            specs=specs,
        )

    @app.route("/categorie/<slug>")
    def category_view(slug):
        category = Category.query.filter_by(slug=slug, active=True).first_or_404()
        page = request.args.get("page", 1, type=int)
        sort = request.args.get("sort", "newest")
        query = Product.query.filter_by(categorie_id=category.id, statut="visible")

        if sort == "price_asc":
            query = query.order_by(Product.prix.asc())
        elif sort == "price_desc":
            query = query.order_by(Product.prix.desc())
        else:
            query = query.order_by(Product.date_creation.desc())

        products = query.paginate(page=page, per_page=app.config["ITEMS_PER_PAGE"], error_out=False)
        return render_template("public/category.html", category=category, products=products, current_sort=sort)

    @app.route("/recherche")
    def search():
        q = request.args.get("q", "").strip()
        page = request.args.get("page", 1, type=int)
        products = None

        if q:
            search_term = f"%{q}%"
            products = Product.query.filter(
                Product.statut == "visible",
                db.or_(
                    Product.nom_fr.ilike(search_term),
                    Product.nom_ar.ilike(search_term),
                    Product.reference.ilike(search_term),
                    Product.marque.ilike(search_term),
                    Product.description_fr.ilike(search_term),
                )
            ).order_by(Product.nb_vues.desc()).paginate(
                page=page, per_page=app.config["ITEMS_PER_PAGE"], error_out=False
            )

        return render_template("public/search.html", query=q, products=products)

    @app.route("/a-propos")
    def about():
        return render_template("public/about.html")

    @app.route("/contact")
    def contact():
        return render_template("public/contact.html")

    # ──────────────────────────────────────────
    #  Language Switch
    # ──────────────────────────────────────────
    @app.route("/lang/<code>")
    def set_lang(code):
        if code in app.config["SUPPORTED_LANGUAGES"]:
            session["lang"] = code
            session.permanent = True
        return redirect(request.referrer or url_for("index"))

    # ──────────────────────────────────────────
    #  AJAX: Track WhatsApp Click
    # ──────────────────────────────────────────
    @app.route("/api/track/whatsapp/<int:product_id>", methods=["POST"])
    def track_whatsapp(product_id):
        product = Product.query.get(product_id)
        if product:
            product.nb_clics_whatsapp += 1
            click = WhatsAppClick(produit_id=product_id, lang=session.get("lang", "fr"))
            db.session.add(click)
            db.session.commit()
        return jsonify({"ok": True})

    # ──────────────────────────────────────────
    #  Static file serving for uploads
    # ──────────────────────────────────────────
    @app.route("/uploads/<path:filename>")
    def uploaded_file(filename):
        return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

    # ═══════════════════════════════════════════
    #  ADMIN ROUTES
    # ═══════════════════════════════════════════

    @app.route("/admin/login", methods=["GET", "POST"])
    @limiter.limit("10 per minute")
    def admin_login():
        if current_user.is_authenticated:
            return redirect(url_for("admin_dashboard"))

        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            admin = Admin.query.filter_by(username=username, actif=True).first()

            if admin and admin.check_password(password):
                admin.derniere_connexion = datetime.utcnow()
                db.session.commit()
                login_user(admin, remember=False)
                session.permanent = True
                log_action(db, admin.id, "login", f"Connexion depuis {request.remote_addr}")
                db.session.commit()
                flash(f"Bienvenue, {admin.username} !", "success")
                return redirect(url_for("admin_dashboard"))
            else:
                flash("Identifiants incorrects. Veuillez réessayer.", "danger")

        return render_template("admin/login.html")

    @app.route("/admin/logout")
    @login_required
    def admin_logout():
        log_action(db, current_user.id, "logout", "")
        db.session.commit()
        logout_user()
        flash("Vous avez été déconnecté.", "info")
        return redirect(url_for("admin_login"))

    @app.route("/admin/")
    @app.route("/admin")
    @login_required
    def admin_dashboard():
        total_products = Product.query.count()
        visible_products = Product.query.filter_by(statut="visible").count()
        rupture_count = Product.query.filter_by(statut="visible").filter(Product.quantite_stock == 0).count()
        threshold = int(SiteSettings.get("seuil_stock_faible", "5"))
        low_stock = Product.query.filter(
            Product.quantite_stock > 0,
            Product.quantite_stock <= threshold,
            Product.statut == "visible"
        ).all()

        # WA clicks today
        today = datetime.utcnow().date()
        wa_today = WhatsAppClick.query.filter(
            db.func.date(WhatsAppClick.date_heure) == today
        ).count()

        # Top products by views
        top_viewed = Product.query.order_by(Product.nb_vues.desc()).limit(5).all()
        top_wa = Product.query.order_by(Product.nb_clics_whatsapp.desc()).limit(5).all()

        # Chart data: last 30 days
        chart_labels = []
        chart_views = []
        chart_wa = []
        for i in range(29, -1, -1):
            day = datetime.utcnow().date() - timedelta(days=i)
            chart_labels.append(day.strftime("%d/%m"))
            views = Product.query.filter(
                db.func.date(Product.date_modification) == day
            ).count()
            wa = WhatsAppClick.query.filter(
                db.func.date(WhatsAppClick.date_heure) == day
            ).count()
            chart_views.append(views)
            chart_wa.append(wa)

        recent_logs = ActivityLog.query.order_by(ActivityLog.date_heure.desc()).limit(10).all()
        total_categories = Category.query.count()

        # Sales KPIs
        total_sales = Sale.query.count()
        sales_turnover = db.session.query(db.func.sum(Sale.montant_total)).filter(Sale.statut != "annule").scalar() or 0.0
        total_encaisse = db.session.query(db.func.sum(Sale.montant_paye)).filter(Sale.statut != "annule").scalar() or 0.0
        total_creances = max(0.0, sales_turnover - total_encaisse)

        return render_template(
            "admin/dashboard.html",
            total_products=total_products,
            visible_products=visible_products,
            rupture_count=rupture_count,
            low_stock=low_stock,
            wa_today=wa_today,
            top_viewed=top_viewed,
            top_wa=top_wa,
            chart_labels=json.dumps(chart_labels),
            chart_views=json.dumps(chart_views),
            chart_wa=json.dumps(chart_wa),
            recent_logs=recent_logs,
            total_categories=total_categories,
            total_sales=total_sales,
            sales_turnover=sales_turnover,
            total_encaisse=total_encaisse,
            total_creances=total_creances,
        )

    # ──────────────────────────────────────────
    #  Products CRUD
    # ──────────────────────────────────────────
    @app.route("/admin/produits")
    @login_required
    def admin_products():
        page = request.args.get("page", 1, type=int)
        search = request.args.get("q", "")
        cat_id = request.args.get("cat", type=int)
        status = request.args.get("statut", "")
        etat = request.args.get("etat", "")

        query = Product.query

        if search:
            query = query.filter(
                db.or_(
                    Product.nom_fr.ilike(f"%{search}%"),
                    Product.reference.ilike(f"%{search}%"),
                    Product.marque.ilike(f"%{search}%"),
                )
            )
        if cat_id:
            query = query.filter_by(categorie_id=cat_id)
        if status:
            query = query.filter_by(statut=status)
        if etat:
            query = query.filter_by(etat=etat)

        products = query.order_by(Product.date_modification.desc()).paginate(
            page=page, per_page=25, error_out=False
        )
        categories = Category.query.order_by(Category.ordre).all()

        return render_template(
            "admin/products/list.html",
            products=products,
            categories=categories,
            search=search,
            current_cat=cat_id,
            current_status=status,
            current_etat=etat,
        )

    @app.route("/admin/produits/nouveau", methods=["GET", "POST"])
    @login_required
    def admin_product_new():
        if not current_user.can_edit:
            abort(403)
        categories = Category.query.order_by(Category.ordre).all()

        if request.method == "POST":
            product, error = _save_product_from_form(None)
            if error:
                flash(error, "danger")
            else:
                log_action(db, current_user.id, "create_product", f"Produit #{product.id} — {product.reference}")
                db.session.commit()
                flash(f"Produit '{product.nom_fr}' créé avec succès !", "success")
                return redirect(url_for("admin_products"))

        return render_template("admin/products/form.html", product=None, categories=categories)

    @app.route("/admin/produits/<int:product_id>/edit", methods=["GET", "POST"])
    @login_required
    def admin_product_edit(product_id):
        if not current_user.can_edit:
            abort(403)
        product = Product.query.get_or_404(product_id)
        categories = Category.query.order_by(Category.ordre).all()

        if request.method == "POST":
            updated_product, error = _save_product_from_form(product)
            if error:
                flash(error, "danger")
            else:
                log_action(db, current_user.id, "edit_product", f"Produit #{product.id} — {product.reference}")
                db.session.commit()
                flash("Produit mis à jour !", "success")
                return redirect(url_for("admin_products"))

        return render_template("admin/products/form.html", product=product, categories=categories)

    @app.route("/admin/produits/<int:product_id>/delete", methods=["POST"])
    @login_required
    def admin_product_delete(product_id):
        if not current_user.can_edit:
            abort(403)
        product = Product.query.get_or_404(product_id)

        # Delete associated images
        for img in product.images:
            delete_product_image(img.chemin_fichier)

        ref = product.reference
        log_action(db, current_user.id, "delete_product", f"Produit #{product_id} — {ref}")
        db.session.delete(product)
        db.session.commit()
        flash(f"Produit {ref} supprimé.", "success")
        return redirect(url_for("admin_products"))

    @app.route("/admin/produits/<int:product_id>/duplicate", methods=["POST"])
    @login_required
    def admin_product_duplicate(product_id):
        if not current_user.can_edit:
            abort(403)
        original = Product.query.get_or_404(product_id)
        new_ref = f"{original.reference}-COPY"
        count = 1
        while Product.query.filter_by(reference=new_ref).first():
            new_ref = f"{original.reference}-COPY{count}"
            count += 1

        dup = Product(
            reference=new_ref,
            nom_fr=f"{original.nom_fr} (Copie)",
            nom_ar=f"{original.nom_ar} (نسخة)",
            description_fr=original.description_fr,
            description_ar=original.description_ar,
            specs=original.specs,
            categorie_id=original.categorie_id,
            marque=original.marque,
            prix=original.prix,
            prix_promo=original.prix_promo,
            quantite_stock=original.quantite_stock,
            etat=original.etat,
            statut="brouillon",
        )
        db.session.add(dup)
        log_action(db, current_user.id, "duplicate_product", f"Copie de #{product_id} → {new_ref}")
        db.session.commit()
        flash(f"Produit dupliqué : {new_ref}", "success")
        return redirect(url_for("admin_product_edit", product_id=dup.id))

    @app.route("/admin/produits/<int:product_id>/toggle", methods=["POST"])
    @login_required
    def admin_product_toggle(product_id):
        if not current_user.can_edit:
            abort(403)
        product = Product.query.get_or_404(product_id)
        product.statut = "masque" if product.statut == "visible" else "visible"
        db.session.commit()
        return jsonify({"status": product.statut})

    @app.route("/admin/produits/<int:product_id>/images/delete/<int:image_id>", methods=["POST"])
    @login_required
    def admin_delete_image(product_id, image_id):
        if not current_user.can_edit:
            abort(403)
        img = ProductImage.query.filter_by(id=image_id, produit_id=product_id).first_or_404()
        delete_product_image(img.chemin_fichier)
        db.session.delete(img)
        db.session.commit()
        return jsonify({"ok": True})

    @app.route("/admin/produits/stock-inline", methods=["POST"])
    @login_required
    def admin_update_stock_inline():
        if not current_user.can_edit:
            abort(403)
        product_id = request.json.get("id")
        qty = request.json.get("qty", 0)
        product = Product.query.get(product_id)
        if product:
            product.quantite_stock = max(0, int(qty))
            db.session.commit()
            return jsonify({"ok": True, "qty": product.quantite_stock})
        return jsonify({"ok": False}), 404

    # ──────────────────────────────────────────
    #  Categories CRUD
    # ──────────────────────────────────────────
    @app.route("/admin/categories", methods=["GET", "POST"])
    @login_required
    def admin_categories():
        if not current_user.can_edit:
            abort(403)

        if request.method == "POST":
            action = request.form.get("action")

            if action == "create":
                nom_fr = request.form.get("nom_fr", "").strip()
                nom_ar = request.form.get("nom_ar", "").strip()
                icone = request.form.get("icone", "bi-box").strip()
                ordre = request.form.get("ordre", 99, type=int)
                slug = slugify(nom_fr)
                count = 1
                base_slug = slug
                while Category.query.filter_by(slug=slug).first():
                    slug = f"{base_slug}-{count}"
                    count += 1

                cat = Category(nom_fr=nom_fr, nom_ar=nom_ar, slug=slug, icone=icone, ordre=ordre)
                db.session.add(cat)
                log_action(db, current_user.id, "create_category", nom_fr)
                db.session.commit()
                flash("Catégorie créée.", "success")

            elif action == "edit":
                cat_id = request.form.get("cat_id", type=int)
                cat = Category.query.get_or_404(cat_id)
                cat.nom_fr = request.form.get("nom_fr", cat.nom_fr).strip()
                cat.nom_ar = request.form.get("nom_ar", cat.nom_ar).strip()
                cat.icone = request.form.get("icone", cat.icone).strip()
                cat.ordre = request.form.get("ordre", cat.ordre, type=int)
                cat.active = bool(request.form.get("active"))
                log_action(db, current_user.id, "edit_category", cat.nom_fr)
                db.session.commit()
                flash("Catégorie mise à jour.", "success")

            elif action == "delete":
                cat_id = request.form.get("cat_id", type=int)
                cat = Category.query.get_or_404(cat_id)
                if Product.query.filter_by(categorie_id=cat.id).count() > 0:
                    flash("Impossible de supprimer : des produits sont liés à cette catégorie.", "danger")
                else:
                    log_action(db, current_user.id, "delete_category", cat.nom_fr)
                    db.session.delete(cat)
                    db.session.commit()
                    flash("Catégorie supprimée.", "success")

            return redirect(url_for("admin_categories"))

        categories = Category.query.order_by(Category.ordre).all()
        cat_product_counts = {
            c.id: Product.query.filter_by(categorie_id=c.id).count()
            for c in categories
        }
        return render_template(
            "admin/categories.html",
            categories=categories,
            cat_product_counts=cat_product_counts,
        )

    # ──────────────────────────────────────────
    #  Settings
    # ──────────────────────────────────────────
    @app.route("/admin/parametres", methods=["GET", "POST"])
    @login_required
    def admin_settings():
        if request.method == "POST":
            keys_to_save = [
                "nom_boutique", "slogan_fr", "slogan_ar", "whatsapp_principal",
                "whatsapp_secondaire", "email", "telephone", "adresse_fr", "adresse_ar",
                "horaires_fr", "horaires_ar", "facebook", "instagram",
                "message_whatsapp_fr", "message_whatsapp_ar",
                "couleur_primaire", "couleur_secondaire", "devise",
                "a_propos_fr", "a_propos_ar", "seuil_stock_faible",
                "mode_maintenance", "afficher_prix", "google_maps_embed",
            ]
            for key in keys_to_save:
                val = request.form.get(key, "")
                if key == "mode_maintenance":
                    val = "1" if request.form.get("mode_maintenance") else "0"
                if key == "afficher_prix":
                    val = "1" if request.form.get("afficher_prix") else "0"
                SiteSettings.set(key, val)

            # Logo upload
            if "logo" in request.files and request.files["logo"].filename:
                from utils.image_handler import save_product_image, allowed_file
                logo_file = request.files["logo"]
                if allowed_file(logo_file.filename):
                    from PIL import Image
                    import uuid
                    logo_filename = f"logo_{uuid.uuid4().hex[:8]}.webp"
                    logo_path = os.path.join(app.config["UPLOAD_FOLDER"], logo_filename)
                    img = Image.open(logo_file.stream).convert("RGB")
                    img.thumbnail((400, 200))
                    img.save(logo_path, "WEBP", quality=90)
                    SiteSettings.set("logo", logo_filename)

            log_action(db, current_user.id, "update_settings", "Paramètres généraux mis à jour")
            db.session.commit()
            flash("Paramètres enregistrés !", "success")
            return redirect(url_for("admin_settings"))

        settings = SiteSettings.get_all_dict()
        return render_template("admin/settings.html", settings=settings)

    # ──────────────────────────────────────────
    #  Users
    # ──────────────────────────────────────────
    @app.route("/admin/utilisateurs", methods=["GET", "POST"])
    @login_required
    def admin_users():
        if not current_user.is_superadmin:
            abort(403)

        if request.method == "POST":
            action = request.form.get("action")

            if action == "create":
                username = request.form.get("username", "").strip()
                email = request.form.get("email", "").strip()
                password = request.form.get("password", "")
                role = request.form.get("role", "manager")

                if Admin.query.filter_by(username=username).first():
                    flash("Ce nom d'utilisateur existe déjà.", "danger")
                else:
                    new_admin = Admin(username=username, email=email, role=role)
                    new_admin.set_password(password)
                    db.session.add(new_admin)
                    log_action(db, current_user.id, "create_user", username)
                    db.session.commit()
                    flash(f"Compte '{username}' créé.", "success")

            elif action == "toggle":
                user_id = request.form.get("user_id", type=int)
                user = Admin.query.get_or_404(user_id)
                if user.id != current_user.id:
                    user.actif = not user.actif
                    db.session.commit()
                    flash("Statut mis à jour.", "success")

            elif action == "delete":
                user_id = request.form.get("user_id", type=int)
                user = Admin.query.get(user_id)
                if user and user.id != current_user.id:
                    log_action(db, current_user.id, "delete_user", user.username)
                    db.session.delete(user)
                    db.session.commit()
                    flash("Compte supprimé.", "success")

            elif action == "change_password":
                user_id = request.form.get("user_id", type=int)
                new_pwd = request.form.get("new_password", "")
                user = Admin.query.get(user_id)
                if user and (user.id == current_user.id or current_user.is_superadmin):
                    user.set_password(new_pwd)
                    db.session.commit()
                    flash("Mot de passe modifié.", "success")

            return redirect(url_for("admin_users"))

        admins = Admin.query.order_by(Admin.date_creation).all()
        return render_template("admin/users.html", admins=admins)

    # ──────────────────────────────────────────
    #  Logs
    # ──────────────────────────────────────────
    @app.route("/admin/logs")
    @login_required
    def admin_logs():
        page = request.args.get("page", 1, type=int)
        logs = ActivityLog.query.order_by(ActivityLog.date_heure.desc()).paginate(
            page=page, per_page=50, error_out=False
        )
        return render_template("admin/logs.html", logs=logs)

    # ──────────────────────────────────────────
    #  Export / Import Excel
    # ──────────────────────────────────────────
    @app.route("/admin/export/excel")
    @login_required
    def admin_export_excel():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produits"

        headers = [
            "ID", "Référence", "Nom FR", "Nom AR", "Catégorie",
            "Marque", "Prix", "Prix Promo", "Stock", "État",
            "Statut", "Vues", "Clics WA", "Date Création"
        ]
        ws.append(headers)

        products = Product.query.order_by(Product.id).all()
        for p in products:
            cat_name = p.categorie.nom_fr if p.categorie else ""
            ws.append([
                p.id, p.reference, p.nom_fr, p.nom_ar, cat_name,
                p.marque, p.prix, p.prix_promo or "", p.quantite_stock,
                p.etat, p.statut, p.nb_vues, p.nb_clics_whatsapp,
                p.date_creation.strftime("%Y-%m-%d %H:%M") if p.date_creation else ""
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"electro_produits_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    @app.route("/admin/import/excel", methods=["POST"])
    @login_required
    def admin_import_excel():
        if not current_user.can_edit:
            abort(403)

        file = request.files.get("excel_file")
        if not file or not file.filename.endswith((".xlsx", ".xls")):
            flash("Fichier Excel invalide.", "danger")
            return redirect(url_for("admin_products"))

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            imported = 0
            updated = 0

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if not row[1]:  # Reference required
                    continue

                ref = str(row[1]).strip()
                nom_fr = str(row[2] or "").strip()
                nom_ar = str(row[3] or "").strip()
                prix = float(row[6]) if row[6] else 0
                prix_promo = float(row[7]) if row[7] else None
                stock = int(row[8]) if row[8] is not None else 0
                etat = str(row[9] or "neuf").strip()

                existing = Product.query.filter_by(reference=ref).first()
                if existing:
                    existing.nom_fr = nom_fr or existing.nom_fr
                    existing.nom_ar = nom_ar or existing.nom_ar
                    existing.prix = prix or existing.prix
                    existing.prix_promo = prix_promo
                    existing.quantite_stock = stock
                    existing.etat = etat
                    updated += 1
                else:
                    product = Product(
                        reference=ref, nom_fr=nom_fr or ref,
                        nom_ar=nom_ar or ref, prix=prix,
                        prix_promo=prix_promo, quantite_stock=stock, etat=etat,
                        statut="brouillon",
                    )
                    db.session.add(product)
                    imported += 1

            log_action(db, current_user.id, "import_excel", f"{imported} importés, {updated} mis à jour")
            db.session.commit()
            flash(f"Import réussi : {imported} produits créés, {updated} mis à jour.", "success")

        except Exception as e:
            flash(f"Erreur lors de l'import : {e}", "danger")

        return redirect(url_for("admin_products"))

    # ──────────────────────────────────────────
    #  Sales Management (Ventes WhatsApp)
    # ──────────────────────────────────────────
    @app.route("/admin/ventes")
    @login_required
    def admin_sales():
        page = request.args.get("page", 1, type=int)
        search = request.args.get("q", "").strip()
        statut_paiement = request.args.get("statut_paiement", "")
        statut = request.args.get("statut", "")

        query = Sale.query

        if search:
            query = query.filter(
                db.or_(
                    Sale.nom_client.ilike(f"%{search}%"),
                    Sale.reference.ilike(f"%{search}%"),
                    Sale.telephone_client.ilike(f"%{search}%"),
                )
            )
        if statut_paiement:
            query = query.filter_by(statut_paiement=statut_paiement)
        if statut:
            query = query.filter_by(statut=statut)

        sales = query.order_by(Sale.date_vente.desc()).paginate(
            page=page, per_page=20, error_out=False
        )

        return render_template(
            "admin/sales/list.html",
            sales=sales,
            search=search,
            current_payment_status=statut_paiement,
            current_status=statut,
        )

    @app.route("/admin/ventes/nouvelle", methods=["GET", "POST"])
    @login_required
    def admin_sale_new():
        if not current_user.can_edit:
            abort(403)
        
        products = Product.query.order_by(Product.reference).all()
        products_list = []
        for p in products:
            products_list.append({
                "id": p.id,
                "ref": p.reference,
                "name": p.nom_fr,
                "price": p.prix_actuel,
                "stock": p.quantite_stock
            })
        
        products_json = json.dumps(products_list, ensure_ascii=False)

        if request.method == "POST":
            try:
                nom_client = request.form.get("nom_client", "").strip()
                telephone_client = request.form.get("telephone_client", "").strip()
                adresse_client = request.form.get("adresse_client", "").strip()
                commentaire = request.form.get("commentaire", "").strip()
                
                date_vente_str = request.form.get("date_vente")
                if date_vente_str:
                    date_vente = datetime.strptime(date_vente_str, "%Y-%m-%dT%H:%M")
                else:
                    date_vente = datetime.utcnow()

                today_str = datetime.utcnow().strftime("%Y%m%d")
                last_sale = Sale.query.filter(Sale.reference.like(f"VNT-{today_str}-%")).order_by(Sale.id.desc()).first()
                if last_sale:
                    last_num = int(last_sale.reference.split("-")[-1])
                    new_num = last_num + 1
                else:
                    new_num = 1
                reference = f"VNT-{today_str}-{new_num:04d}"

                sale = Sale(
                    reference=reference,
                    date_vente=date_vente,
                    nom_client=nom_client,
                    telephone_client=telephone_client,
                    adresse_client=adresse_client,
                    commentaire=commentaire,
                    statut="complete",
                    montant_total=0.0,
                    montant_paye=0.0,
                    statut_paiement="non_paye"
                )
                db.session.add(sale)
                db.session.flush()

                total_amount = 0.0
                item_indices = set()
                for key in request.form.keys():
                    if key.startswith("items[") and "]" in key:
                        idx = key.split("]")[0].split("[")[1]
                        item_indices.add(idx)

                for idx in item_indices:
                    p_id = request.form.get(f"items[{idx}][product_id]", type=int)
                    qty = request.form.get(f"items[{idx}][quantite]", type=int)
                    unit_price = request.form.get(f"items[{idx}][prix_unitaire]", type=float)

                    if p_id and qty and unit_price is not None:
                        prod = Product.query.get(p_id)
                        if prod:
                            sub_total = qty * unit_price
                            total_amount += sub_total
                            
                            sale_item = SaleItem(
                                sale_id=sale.id,
                                produit_id=prod.id,
                                reference_produit=prod.reference,
                                nom_produit=prod.nom_fr,
                                quantite=qty,
                                prix_unitaire=unit_price,
                                sous_total=sub_total
                            )
                            db.session.add(sale_item)
                            
                            prod.quantite_stock = max(0, prod.quantite_stock - qty)
                            log_action(db, current_user.id, "stock_decrement", f"Stock de {prod.reference} décrémenté de {qty} (Vente {reference})")

                sale.montant_total = total_amount

                initial_pay = request.form.get("montant_paye", 0.0, type=float)
                if initial_pay > 0:
                    mode = request.form.get("mode_paiement", "Espèces")
                    ref_pay = request.form.get("reference_paiement", "").strip()
                    notes_pay = request.form.get("notes_paiement", "").strip()
                    
                    preuve_file = request.files.get("preuve_paiement")
                    preuve_path = None
                    if preuve_file and preuve_file.filename:
                        preuve_path = save_payment_proof(preuve_file, reference)

                    payment = Payment(
                        sale_id=sale.id,
                        date_paiement=date_vente,
                        montant=initial_pay,
                        mode_paiement=mode,
                        reference_paiement=ref_pay,
                        preuve_paiement=preuve_path,
                        notes=notes_pay
                    )
                    db.session.add(payment)
                    db.session.flush()
                
                sale.recalculate_totals()
                
                log_action(db, current_user.id, "create_sale", f"Vente {reference} créée pour client {nom_client} d'un total de {total_amount} DH")
                db.session.commit()
                flash(f"Vente '{reference}' enregistrée avec succès !", "success")
                return redirect(url_for("admin_sales"))

            except Exception as e:
                db.session.rollback()
                flash(f"Erreur lors de la création de la vente : {e}", "danger")

        return render_template("admin/sales/form.html", products_json=products_json)

    @app.route("/admin/ventes/<int:sale_id>", methods=["GET", "POST"])
    @login_required
    def admin_sale_detail(sale_id):
        sale = Sale.query.get_or_404(sale_id)

        if request.method == "POST":
            if not current_user.can_edit:
                abort(403)
            try:
                montant = request.form.get("montant", type=float)
                if not montant or montant <= 0:
                    flash("Le montant du paiement doit être supérieur à 0.", "danger")
                elif montant > sale.reste_a_payer:
                    flash(f"Le montant ne peut pas dépasser le reste à payer ({sale.reste_a_payer} DH).", "danger")
                else:
                    mode = request.form.get("mode_paiement", "Espèces")
                    ref_pay = request.form.get("reference_paiement", "").strip()
                    notes = request.form.get("notes", "").strip()
                    
                    preuve_file = request.files.get("preuve_paiement")
                    preuve_path = None
                    if preuve_file and preuve_file.filename:
                        preuve_path = save_payment_proof(preuve_file, sale.reference)

                    payment = Payment(
                        sale_id=sale.id,
                        date_paiement=datetime.utcnow(),
                        montant=montant,
                        mode_paiement=mode,
                        reference_paiement=ref_pay,
                        preuve_paiement=preuve_path,
                        notes=notes
                    )
                    db.session.add(payment)
                    db.session.flush()
                    
                    sale.recalculate_totals()
                    
                    log_action(db, current_user.id, "add_payment", f"Paiement de {montant} DH ajouté pour la vente {sale.reference}")
                    db.session.commit()
                    flash(f"Paiement de {montant} DH enregistré.", "success")
                    return redirect(url_for("admin_sale_detail", sale_id=sale.id))
            except Exception as e:
                db.session.rollback()
                flash(f"Erreur lors de l'enregistrement du paiement : {e}", "danger")

        return render_template("admin/sales/detail.html", sale=sale)

    @app.route("/admin/ventes/<int:sale_id>/status", methods=["POST"])
    @login_required
    def admin_sale_status(sale_id):
        if not current_user.can_edit:
            abort(403)
        sale = Sale.query.get_or_404(sale_id)
        old_status = sale.statut
        new_status = request.form.get("statut", "").strip()

        if new_status in ["complete", "en_attente", "annule"] and new_status != old_status:
            try:
                if new_status == "annule":
                    for item in sale.items:
                        if item.produit:
                            item.produit.quantite_stock += item.quantite
                            log_action(db, current_user.id, "stock_increment", f"Stock de {item.reference_produit} incrémenté de {item.quantite} (Vente {sale.reference} Annulée)")
                elif old_status == "annule":
                    for item in sale.items:
                        if item.produit:
                            item.produit.quantite_stock = max(0, item.produit.quantite_stock - item.quantite)
                            log_action(db, current_user.id, "stock_decrement", f"Stock de {item.reference_produit} décrémenté de {item.quantite} (Vente {sale.reference} Réactivée)")

                sale.statut = new_status
                log_action(db, current_user.id, "update_sale_status", f"Statut de la vente {sale.reference} changé de {old_status} à {new_status}")
                db.session.commit()
                flash("Statut de la vente mis à jour.", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Erreur lors de la mise à jour du statut : {e}", "danger")

        return redirect(url_for("admin_sale_detail", sale_id=sale.id))

    @app.route("/admin/ventes/<int:sale_id>/paiement/<int:payment_id>/supprimer", methods=["POST"])
    @login_required
    def admin_payment_delete(sale_id, payment_id):
        if not current_user.can_edit:
            abort(403)
        sale = Sale.query.get_or_404(sale_id)
        payment = Payment.query.filter_by(id=payment_id, sale_id=sale_id).first_or_404()

        try:
            if payment.preuve_paiement:
                upload_dir = current_app.config["UPLOAD_FOLDER"]
                file_path = os.path.join(upload_dir, payment.preuve_paiement)
                if os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                    except Exception as ex:
                        current_app.logger.warning(f"Could not delete proof file {file_path}: {ex}")

            montant = payment.montant
            db.session.delete(payment)
            db.session.flush()
            
            sale.recalculate_totals()
            
            log_action(db, current_user.id, "delete_payment", f"Paiement de {montant} DH supprimé pour la vente {sale.reference}")
            db.session.commit()
            flash("Paiement supprimé avec succès.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erreur lors de la suppression du paiement : {e}", "danger")

        return redirect(url_for("admin_sale_detail", sale_id=sale.id))

    @app.route("/admin/ventes/<int:sale_id>/supprimer", methods=["POST"])
    @login_required
    def admin_sale_delete(sale_id):
        if not current_user.can_edit:
            abort(403)
        sale = Sale.query.get_or_404(sale_id)

        try:
            if sale.statut != "annule":
                for item in sale.items:
                    if item.produit:
                        item.produit.quantite_stock += item.quantite
                        log_action(db, current_user.id, "stock_increment", f"Stock de {item.reference_produit} incrémenté de {item.quantite} (Vente {sale.reference} Supprimée)")

            for payment in sale.payments:
                if payment.preuve_paiement:
                    upload_dir = current_app.config["UPLOAD_FOLDER"]
                    file_path = os.path.join(upload_dir, payment.preuve_paiement)
                    if os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                        except Exception as ex:
                            current_app.logger.warning(f"Could not delete proof file {file_path}: {ex}")

            ref = sale.reference
            db.session.delete(sale)
            db.session.commit()
            flash(f"Vente '{ref}' supprimée.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erreur lors de la suppression de la vente : {e}", "danger")

        return redirect(url_for("admin_sales"))

    # ──────────────────────────────────────────
    #  Internal helper: save product from form
    # ──────────────────────────────────────────
    def _save_product_from_form(product):
        try:
            nom_fr = request.form.get("nom_fr", "").strip()
            if not nom_fr:
                return None, "Le nom en français est requis."

            reference = request.form.get("reference", "").strip()
            if not reference:
                return None, "La référence est requise."

            # Check unique reference
            existing = Product.query.filter_by(reference=reference).first()
            if existing and (product is None or existing.id != product.id):
                return None, f"La référence '{reference}' existe déjà."

            prix = request.form.get("prix", 0, type=float)

            # Build specs dict from dynamic form
            spec_keys = request.form.getlist("spec_key[]")
            spec_vals = request.form.getlist("spec_val[]")
            specs_dict = {k: v for k, v in zip(spec_keys, spec_vals) if k.strip()}

            promo_debut = None
            promo_fin = None
            if request.form.get("promo_debut"):
                promo_debut = datetime.strptime(request.form.get("promo_debut"), "%Y-%m-%d")
            if request.form.get("promo_fin"):
                promo_fin = datetime.strptime(request.form.get("promo_fin"), "%Y-%m-%d")

            if product is None:
                product = Product()
                db.session.add(product)

            product.reference = reference
            product.nom_fr = nom_fr
            product.nom_ar = request.form.get("nom_ar", "").strip()
            product.description_fr = request.form.get("description_fr", "").strip()
            product.description_ar = request.form.get("description_ar", "").strip()
            product.categorie_id = request.form.get("categorie_id", type=int) or None
            product.marque = request.form.get("marque", "").strip()
            product.prix = prix
            product.prix_promo = request.form.get("prix_promo", type=float) or None
            product.promo_debut = promo_debut
            product.promo_fin = promo_fin
            product.quantite_stock = request.form.get("quantite_stock", 0, type=int)
            product.etat = request.form.get("etat", "neuf")
            product.statut = request.form.get("statut", "visible")
            product.en_vedette = bool(request.form.get("en_vedette"))
            product.specs = json.dumps(specs_dict, ensure_ascii=False)
            product.date_modification = datetime.utcnow()

            db.session.flush()

            # Handle image uploads
            files = request.files.getlist("images[]")
            existing_order = product.images.count()
            for file in files:
                if file and file.filename:
                    path = save_product_image(file, product.reference)
                    if path:
                        img = ProductImage(produit_id=product.id, chemin_fichier=path, ordre=existing_order)
                        db.session.add(img)
                        existing_order += 1

            return product, None

        except Exception as e:
            return None, str(e)

    # ──────────────────────────────────────────
    #  Error Handlers
    # ──────────────────────────────────────────
    @app.errorhandler(404)
    def not_found(e):
        return render_template("errors/404.html"), 404

    @app.errorhandler(403)
    def forbidden(e):
        return render_template("errors/403.html"), 403

    @app.errorhandler(413)
    def too_large(e):
        flash("Fichier trop volumineux (max 5 Mo).", "danger")
        return redirect(request.referrer or url_for("index"))

    return app


# ─────────────────────────────────────────────────────────
#  Entry Point
# ─────────────────────────────────────────────────────────
app = create_app()

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(debug=True, host="0.0.0.0", port=5000)
